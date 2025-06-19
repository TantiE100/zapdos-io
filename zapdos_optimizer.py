import datetime, io, os
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from pulp import LpProblem, LpMinimize, LpVariable, lpSum, PULP_CBC_CMD
import base64
import plotly.graph_objects as go
import plotly.express as px

def get_base64_image(image_path):
    """Convierte una imagen a base64 para usar en CSS"""
    try:
        with open(image_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    except:
        return None

tmpl_dir     = "templates"
diario_input = os.path.join(tmpl_dir, "diario_base.xlsx")
plan_tmpl    = os.path.join(tmpl_dir, "plan_base.xlsx")

st.set_page_config(page_title="Zapdos", page_icon="favicon.png", layout="centered")

bg_image = get_base64_image("images/bg.jpg")
if bg_image:
    st.markdown(f"""
    <style>
        .stApp {{
            background: linear-gradient(rgba(14, 17, 23, 0.75), rgba(14, 17, 23, 0.75)), 
                        url('data:image/jpeg;base64,{bg_image}') center center;
            background-size: cover;
            background-attachment: fixed;
        }}
    </style>
    """, unsafe_allow_html=True)

st.markdown("""
<style>
     /* Mejorar legibilidad del texto */
     .main .block-container {
         background-color: rgba(14, 17, 23, 0.85);
         border-radius: 15px;
         padding: 2rem;
         margin-top: 2rem;
         backdrop-filter: blur(10px);
         box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
     }
    
    /* Estilo para títulos */
    h1, h2, h3 {
        color: #FFD700 !important;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.8);
    }
    
    /* Mejorar contraste en inputs */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input {
        background-color: rgba(38, 39, 48, 0.9) !important;
        color: white !important;
        border: 1px solid #FFD700 !important;
    }
    
    /* Botones con estilo Team Instinct */
    .stButton > button {
        background: linear-gradient(45deg, #FFD700, #FFA500) !important;
        color: black !important;
        font-weight: bold !important;
        border: none !important;
        box-shadow: 0 4px 8px rgba(255, 215, 0, 0.3) !important;
    }
    
    .stButton > button:hover {
        background: linear-gradient(45deg, #FFA500, #FFD700) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 12px rgba(255, 215, 0, 0.5) !important;
    }
</style>
""", unsafe_allow_html=True)

st.title("⚡ Zapdos — Plan Diario")

# 1. Descarga plantilla diaria
st.header("1. Descarga plantilla de 24 horas, y llena los datos")
with open(diario_input, "rb") as f:
    st.download_button("📥 Descargar plantilla 24 h", data=f,
                       file_name="plantilla_diaria.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       key="download_template")

# 2. Carga de datos
st.header("2. Sube tu archivo horario diario (.xlsx / .csv)")

uploaded = st.file_uploader("Archivo de 24 filas (.xlsx)", type=["xlsx"], key="upload_schedule")
if not uploaded:
    st.stop()

try:
    df = pd.read_excel(uploaded)
except Exception:
    st.error("El archivo no es un .xlsx válido.")
    st.stop()

req_cols = [
    "Generación solar disponible (kW)",
    "Tarifa de compra (Bs / kWh)",
    "Tarifa de venta (Bs / kWh)",
    "Consumo base obligatorio (kW)"
]
faltantes = [c for c in req_cols if c not in df.columns]
if faltantes:
    st.error(f"Faltan columnas: {', '.join(faltantes)}. Descarga y usa la plantilla.")
    st.stop()

# Limpieza
solar_col = "Generación solar disponible (kW)"
df[solar_col] = df[solar_col].replace("-", 0).fillna(0).astype(float)
df["Tarifa de compra (Bs / kWh)"]   = df["Tarifa de compra (Bs / kWh)"].astype(float)
df["Tarifa de venta (Bs / kWh)"]    = df["Tarifa de venta (Bs / kWh)"].astype(float)
df["Consumo base obligatorio (kW)"] = df["Consumo base obligatorio (kW)"].astype(float)

st.success("📊 Datos diarios cargados")
st.dataframe(df, width=1000, height=300)

# 3. Parámetros del equipo
st.header("3. Configura tu equipo")
nombre = st.text_input("Nombre del equipo", placeholder="cafetera, refrigerador, etc.", key="input_name")

col1, col2, col3 = st.columns([1,2,1])
with col2:
    st.image("images/potencia-ejemplo.png", width=500, caption="Imagen de ejemplo: Potencia nominal")

pot_w = st.number_input(
    "Potencia nominal (W)",
    min_value=1,
    max_value=20000,
    value=1500,
    key="input_power"
)
pot_kW = pot_w / 1000  # Definir pot_kW

col1, col2, col3 = st.columns([1,2,1])
with col2:
    st.image("images/factura-ejemplo.png", width=500)

pot_contrat = st.number_input(
    "Potencia contratada (kW)",
    min_value=0.1,
    value=7.0,
    max_value=100.0,
    key="input_contract"
)

horas_min = st.number_input(
    "Horas mínimas de funcionamiento",
    min_value=1,
    max_value=24,
    value=3,
    key="input_horas"
)

# 4. Optimización
if st.button("🚀 Calcular plan diario óptimo", key="btn_optimize"):
    if not nombre.strip():
        st.error("Debes ingresar un nombre de equipo antes de optimizar.")
    else:
        # PROGRAMACION LINEAL
        T = range(24)
        Compra = LpVariable.dicts("Compra_kW", T, 0)
        Venta  = LpVariable.dicts("Inyeccion_kW", T, 0)
        ON     = LpVariable.dicts("ON", T, cat="Binary")

        model = LpProblem("ZapdosDiario", LpMinimize)
        model += lpSum(df["Tarifa de compra (Bs / kWh)"][t] * Compra[t] -
                    df["Tarifa de venta (Bs / kWh)"][t] * Venta[t] for t in T)

        for t in T:
            model += (Compra[t] + df[solar_col][t] ==
                    df["Consumo base obligatorio (kW)"][t] + pot_kW * ON[t] + Venta[t])
            model += Compra[t] <= pot_contrat
        model += lpSum(ON[t] for t in T) == horas_min
        model.solve(PULP_CBC_CMD(msg=False))

        # 5. Plan + métricas
        costo_base = (df["Tarifa de compra (Bs / kWh)"] *
                    df["Consumo base obligatorio (kW)"]).sum()
        costo_opt  = 0
        filas = []
        for t in T:
            comp = Compra[t].value(); iny = Venta[t].value()
            costo_opt += (df["Tarifa de compra (Bs / kWh)"][t] * comp -
                        df["Tarifa de venta (Bs / kWh)"][t] * iny)
            ahorro = costo_base - costo_opt
            pct    = ahorro / costo_base if costo_base else 0
            filas.append((t+1, int(ON[t].value()), comp, iny, ahorro, pct))

        plan = pd.DataFrame(filas, columns=[
            "Intervalo horario (h)", "Estado equipo (0/1)",
            "Potencia comprada a la red (kW)",
            "Potencia inyectada a la red (kW)",
            "Ahorro acumulado (Bs)", "Ahorro vs. Base (%)"
        ])

        # 7. Rellenar plantilla y descargar XLSX
        wb = load_workbook(plan_tmpl)
        ws = wb.active

        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        t = ws.cell(row=1, column=1, value=f"PLAN ÓPTIMO DIARIO - {nombre}".upper())
        t.font = Font(bold=True, size=16)
        t.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        header_fill = PatternFill(start_color='FFE6E6E6', end_color='FFE6E6E6', fill_type='solid')
        header_font = Font(bold=True)
        thin = Side(border_style='thin', color='FF000000')

        for cell in ws.iter_cols(min_row=2, max_row=2, min_col=1, max_col=6):
            for c in cell:
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = Border(bottom=thin)

        start = 3
        
        lavender20 = PatternFill(start_color='FFEFE8FF', end_color='FFEFE8FF', fill_type='solid')
        pink20     = PatternFill(start_color='FFFFF2F2', end_color='FFFFF2F2', fill_type='solid')
        green20    = PatternFill(start_color='FFE8FFE8', end_color='FFE8FFE8', fill_type='solid')
        pink60     = PatternFill(start_color='FFFFC0C0', end_color='FFFFC0C0', fill_type='solid')
        green60    = PatternFill(start_color='FFC0FFC0', end_color='FFC0FFC0', fill_type='solid')

        start = 3
        for r, row in enumerate(plan.itertuples(index=False), start=start):
            hora_txt = f"{row[0]:02d}:00" if row[0] < 24 else "24:00"
            estado   = "ON" if row[1] == 1 else "OFF"

            c_hora = ws.cell(r, 1, hora_txt)
            c_hora.fill = lavender20

            c_status = ws.cell(r, 2, estado)
            c_status.fill = green20 if estado=='ON' else pink20

            c_compra = ws.cell(r, 3, row[2])
            c_compra.number_format = numbers.FORMAT_NUMBER_00
            c_compra.fill = pink60 if row[2] != 0 else green60

            if row[3] == 0:
                c_iny = ws.cell(r, 4, "Sin exposición solar")
                c_iny.fill = pink20
            else:
                c_iny = ws.cell(r, 4, row[3])
                c_iny.number_format = numbers.FORMAT_NUMBER_00

            c_ahorro = ws.cell(r, 5, row[4])
            c_ahorro.number_format = numbers.FORMAT_NUMBER_00

            c_pct = ws.cell(r, 6, row[5])
            c_pct.number_format = "0.00%"

        bd = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows(min_row=start, max_row=start+23, min_col=1, max_col=6):
            for cell in row:
                cell.border = bd

        green_scale = ColorScaleRule(start_type='min', start_color='FFE8FBE8',
                                    mid_type='percentile', mid_value=50, mid_color='FF8FD88F',
                                    end_type='max', end_color='FF008000')
        blue_scale = ColorScaleRule(start_type='min', start_color='FFEAF4FF',
                                    mid_type='percentile', mid_value=50, mid_color='FF7AB8FF',
                                    end_type='max', end_color='FF0066CC')

        ws.conditional_formatting.add(f'D{start}:D{start+23}', green_scale)
        ws.conditional_formatting.add(f'E{start}:E{start+23}', green_scale)
        ws.conditional_formatting.add(f'F{start}:F{start+23}', blue_scale)

        ws.freeze_panes = 'A3'
        if ws.tables:
            ws.tables.clear()

        buf = io.BytesIO(); wb.save(buf)
        fecha = datetime.datetime.now().strftime("%d-%m-%Y")
        file_name = f"plan_optimo_{nombre.lower()}_{horas_min}_{fecha}.xlsx"
        st.download_button("⬇️ Descargar plan (.xlsx)", buf.getvalue(),
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_xlsx_{horas_min}_{fecha}")

        st.success("✅ Plan diario generado")
        
        # ============================================================================
        # 6. DASHBOARD DE OPTIMIZACIÓN
        # ============================================================================
        
        st.markdown("---")
        st.header("📊 Dashboard de Optimización")
        
        # Calcular métricas clave
        ahorro_final = plan["Ahorro acumulado (Bs)"].iloc[-1]
        ahorro_pct_final = plan["Ahorro vs. Base (%)"].iloc[-1] * 100
        
        total_comprado = plan["Potencia comprada a la red (kW)"].sum()
        total_inyectado = plan["Potencia inyectada a la red (kW)"].sum()
        
        horas_funcionamiento = plan["Estado equipo (0/1)"].sum()
        energia_equipo = horas_funcionamiento * pot_kW
        
        costo_sin_equipo = costo_base
        costo_con_equipo = costo_opt
        
        # Tarifa promedio
        tarifa_compra_prom = df["Tarifa de compra (Bs / kWh)"].mean()
        tarifa_venta_prom = df["Tarifa de venta (Bs / kWh)"].mean()
        
        # Ingresos por venta
        ingresos_venta = (plan["Potencia inyectada a la red (kW)"] * 
                         [df["Tarifa de venta (Bs / kWh)"][i] for i in range(24)]).sum()
        
        # Métricas principales en columnas
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                label="💰 Ahorro Total",
                value=f"{ahorro_final:.2f} Bs",
                delta=f"{ahorro_pct_final:.1f}% vs base"
            )
        
        with col2:
            st.metric(
                label="⚡ Energía del Equipo",
                value=f"{energia_equipo:.2f} kWh",
                delta=f"{horas_funcionamiento:.0f} horas ON"
            )
        
        with col3:
            st.metric(
                label="📈 Energía Inyectada",
                value=f"{total_inyectado:.2f} kW",
                delta=f"+{ingresos_venta:.2f} Bs"
            )
        
        with col4:
            st.metric(
                label="📉 Energía Comprada",
                value=f"{total_comprado:.2f} kW",
                delta=f"-{(total_comprado * tarifa_compra_prom):.2f} Bs"
            )
        
        # Gráfico de optimización por horas
        st.subheader("⏰ Optimización por Intervalos Horarios")
        
        # Crear DataFrame para visualización
        vis_data = plan.copy()
        vis_data['Hora'] = [f"{i:02d}:00" for i in range(1, 25)]
        vis_data['Costo Horario'] = [
            df["Tarifa de compra (Bs / kWh)"][i] * plan["Potencia comprada a la red (kW)"].iloc[i] -
            df["Tarifa de venta (Bs / kWh)"][i] * plan["Potencia inyectada a la red (kW)"].iloc[i]
            for i in range(24)
        ]
        
        fig_estado = go.Figure()
        
        # Barras para estado ON/OFF
        colors = ['#2E8B57' if x == 1 else '#DC143C' for x in vis_data['Estado equipo (0/1)']]
        
        fig_estado.add_trace(go.Bar(
            x=vis_data['Hora'],
            y=vis_data['Estado equipo (0/1)'],
            marker_color=colors,
            name='Estado del Equipo',
            text=['ON' if x == 1 else 'OFF' for x in vis_data['Estado equipo (0/1)']],
            textposition='auto',
        ))
        
        fig_estado.update_layout(
            title=f"Estado Óptimo del Equipo: {nombre}",
            xaxis_title="Hora del Día",
            yaxis_title="Estado (0=OFF, 1=ON)",
            showlegend=False,
            height=400,
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(color='white')
        )
        
        st.plotly_chart(fig_estado, use_container_width=True)
        
        # Gráfico de ahorro acumulado
        fig_ahorro = go.Figure()
        
        fig_ahorro.add_trace(go.Scatter(
            x=vis_data['Hora'],
            y=vis_data['Ahorro acumulado (Bs)'],
            mode='lines+markers',
            line=dict(color='#FFD700', width=3),
            marker=dict(size=6, color='#FFA500'),
            name='Ahorro Acumulado',
            fill='tozeroy',
            fillcolor='rgba(255, 215, 0, 0.1)'
        ))
        
        fig_ahorro.update_layout(
            title="Progresión del Ahorro a lo Largo del Día",
            xaxis_title="Hora del Día",
            yaxis_title="Ahorro Acumulado (Bs)",
            showlegend=False,
            height=400,
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(color='white')
        )
        
        st.plotly_chart(fig_ahorro, use_container_width=True)
        
        # Resumen de costos comparativo
        st.subheader("💸 Análisis de Costos")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**🔴 Escenario Sin Optimización:**")
            st.write(f"• Costo base: {costo_sin_equipo:.2f} Bs")
            st.write(f"• Sin uso del equipo")
            st.write(f"• Sin aprovechamiento solar")
            
        with col2:
            st.markdown("**🟢 Escenario Optimizado:**")
            st.write(f"• Costo total: {costo_con_equipo:.2f} Bs")
            st.write(f"• Equipo funcionando {horas_funcionamiento} horas")
            st.write(f"• Energía solar aprovechada: {total_inyectado:.2f} kW")
            
        # Tabla resumen
        st.subheader("📋 Resumen Ejecutivo")
        
        resumen_data = {
            "Métrica": [
                "Ahorro Total (Bs)",
                "Ahorro Porcentual (%)",
                "Horas de Funcionamiento",
                "Energía Consumida por Equipo (kWh)",
                "Energía Comprada Total (kW)",
                "Energía Inyectada Total (kW)",
                "Ingresos por Venta (Bs)",
                "Costo Sin Optimización (Bs)",
                "Costo Con Optimización (Bs)"
            ],
            "Valor": [
                f"{ahorro_final:.2f}",
                f"{ahorro_pct_final:.1f}%",
                f"{horas_funcionamiento:.0f}",
                f"{energia_equipo:.2f}",
                f"{total_comprado:.2f}",
                f"{total_inyectado:.2f}",
                f"{ingresos_venta:.2f}",
                f"{costo_sin_equipo:.2f}",
                f"{costo_con_equipo:.2f}"
            ]
        }
        
        resumen_df = pd.DataFrame(resumen_data)
        st.dataframe(resumen_df, use_container_width=True, hide_index=True)
        
        # Recomendaciones
        st.subheader("💡 Recomendaciones")
        
        if ahorro_final > 0:
            st.success(f"✅ **Excelente optimización**: Estás ahorrando {ahorro_final:.2f} Bs ({ahorro_pct_final:.1f}%) al día")
        else:
            st.warning("⚠️ **Sin ahorro**: Considera ajustar los parámetros o revisar las tarifas")
            
        if total_inyectado > 0:
            st.info(f"☀️ **Energía solar aprovechada**: {total_inyectado:.2f} kW generando {ingresos_venta:.2f} Bs")
        else:
            st.info("🌙 **Sin generación solar**: El plan se optimizó solo con tarifas de compra")
            
        if horas_funcionamiento == horas_min:
            st.info(f"⏱️ **Funcionamiento mínimo**: El equipo opera exactamente {horas_min} horas como se requiere")
        
        # Proyección mensual
        ahorro_mensual = ahorro_final * 30
        ahorro_anual = ahorro_final * 365
        
        st.subheader("📈 Proyección de Ahorro")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Ahorro Diario", f"{ahorro_final:.2f} Bs")
        with col2:
            st.metric("Ahorro Mensual (est.)", f"{ahorro_mensual:.2f} Bs")
        with col3:
            st.metric("Ahorro Anual (est.)", f"{ahorro_anual:.2f} Bs")

st.markdown("---")
st.caption("© 2025 • Zapdos - Investigación Operativa I")